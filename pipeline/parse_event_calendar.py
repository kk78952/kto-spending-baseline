"""
parse_event_calendar.py
───────────────────────
Parses KTO_Nội_Dung_Vận_Hành.xlsx into events.json.

Real sheet layout (discovered from file inspection):
  Row 0 : metadata (report-date stamp, month label)
  Row 1 : actual dates as datetime objects starting at column index 2
  Row 2 : day-of-week labels (2,3,4,5,6,7,CN …)
  Row 3 : REV(M) row in 2024+ sheets  ← col[1] is None  → skip
  Row 3+ (2023) / Row 4+ (2024+):
           event tracks where col[0] = section label or None
                              col[1] = track number (integer)
                              col[2+] = event name at start, None for continuation days

Usage:
  python pipeline/parse_event_calendar.py
  python pipeline/parse_event_calendar.py --xlsx event_calendar/KTO_Nội_Dung_Vận_Hành.xlsx
"""

import argparse
import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dep. Run: pip install openpyxl")

ROOT       = Path(__file__).resolve().parent.parent
XLSX_PATH  = ROOT / "event_calendar" / "KTO_Nội_Dung_Vận_Hành.xlsx"
EVENTS_OUT = ROOT / "event_calendar" / "events.json"

DATE_COL_START = 2   # column index where dates begin (0-indexed)
DATE_ROW_IDX   = 1   # dates live in row index 1 (not 0)
EVENT_ROW_START= 3   # event tracks start at row index 3


def _to_date_str(val) -> str | None:
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    return None


def _is_track_row(row) -> bool:
    """True if this row is an event-track row (col[1] holds a track number)."""
    try:
        v = row[1]
        if v is None:
            return False
        int(str(v).strip())   # raises ValueError if not a number
        return True
    except (ValueError, IndexError, TypeError):
        return False


def parse_sheet(ws, year: int) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < EVENT_ROW_START + 1:
        return []

    # ── Build col_index → date_str from the real date row (index 1)
    date_row = rows[DATE_ROW_IDX]
    col_dates: dict[int, str] = {}
    for ci, val in enumerate(date_row):
        if ci < DATE_COL_START:
            continue
        ds = _to_date_str(val)
        if ds:
            col_dates[ci] = ds

    if not col_dates:
        return []

    max_ci = max(col_dates)

    events = []
    for row in rows[EVENT_ROW_START:]:
        if not _is_track_row(row):
            continue   # skip REV(M) row, header rows, empty rows

        track_num = row[1]

        # Scan across all date columns for event names
        ci = DATE_COL_START
        while ci <= max_ci:
            if ci >= len(row):
                break
            cell_val = row[ci]

            if cell_val is not None and ci in col_dates:
                event_name = str(cell_val).strip()
                if not event_name or event_name.lower() in ("nan", ""):
                    ci += 1
                    continue

                start_date = col_dates[ci]

                # Scan right: Nones = continuation days, next non-None = next event
                end_ci = ci + 1
                while end_ci <= max_ci:
                    if end_ci >= len(row):
                        break
                    if row[end_ci] is not None:
                        break
                    end_ci += 1

                # end_date = last valid date col before the next event cell
                last_ci = end_ci - 1
                while last_ci > ci and last_ci not in col_dates:
                    last_ci -= 1
                end_date = col_dates.get(last_ci, start_date)

                try:
                    track = int(str(track_num).strip())
                except (ValueError, TypeError):
                    track = None

                events.append({
                    "start_date": start_date,
                    "end_date":   end_date,
                    "name":       event_name,
                    "track":      track,
                    "year":       year,
                })
                ci = end_ci   # jump past this event
            else:
                ci += 1

    return events


def parse_xlsx(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    all_events = []

    for sheet_name in wb.sheetnames:
        if not re.match(r"^\d{4}$", sheet_name.strip()):
            continue
        year = int(sheet_name.strip())
        ws   = wb[sheet_name]
        events = parse_sheet(ws, year)
        all_events.extend(events)
        print(f"[parse_event_calendar] sheet {sheet_name}: {len(events)} events")

    # Deduplicate and sort by start date
    seen, unique = set(), []
    for ev in sorted(all_events, key=lambda e: (e["start_date"], e.get("track") or 0)):
        key = (ev["start_date"], ev["name"])
        if key not in seen:
            seen.add(key)
            unique.append(ev)

    return unique


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=str(XLSX_PATH))
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        # Try finding any xlsx in the event_calendar folder
        folder = xlsx_path.parent
        candidates = list(folder.glob("*.xlsx"))
        if candidates:
            xlsx_path = candidates[0]
            print(f"[parse_event_calendar] using {xlsx_path.name}")
        else:
            sys.exit(f"[parse_event_calendar] xlsx not found: {xlsx_path}")

    events = parse_xlsx(xlsx_path)
    print(f"[parse_event_calendar] total unique events: {len(events)}")

    with open(EVENTS_OUT, "w", encoding="utf-8") as f:
        json.dump(events, f, ensure_ascii=False, indent=2)

    print(f"[parse_event_calendar] wrote {EVENTS_OUT}")


if __name__ == "__main__":
    main()
